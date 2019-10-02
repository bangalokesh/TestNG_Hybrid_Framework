import json
import glob
import openpyxl
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook

ROW_INDEX = 1       # keep track of Excel row index
VARIABLES = 7
VALUES = [None] * VARIABLES
VALUES_IDX = 0
VALUES_FASTER = 0
wb = None
FIRST_WRITE = True
SHEET_NAME = "Saved Data"
PATH_TO_FILE = './www/static/saved_data_val.xlsx'

#-------------------------------------------------------

def writeHeader():
    global wb
    global VALUES
    global ROW_INDEX

    sheet = wb[SHEET_NAME]
    VALUES[0]   = "Portal Name"
    VALUES[1]   = "Run ID"
    VALUES[2]   = "Workflow Name"
    VALUES[3]   = "Screen Name"

    pad = 4     # 4 meta values : portal_name,
                # workflow_name, run_id, screen_name

    VALUES[pad + 0]   = "Extract"
    VALUES[pad + 1]   = "Type"
    VALUES[pad + 2]   = "Message"

    for itr in range(VARIABLES):
        temp = chr(65 + itr)    # if itr = 0, temp = A (65 is its ascii value)
        sheet[temp + '1'] = VALUES[itr]

    ROW_INDEX = ROW_INDEX + 1

#------------------------------------------------------

def createNewExcelFile():
    global wb

    wb = Workbook()
    sheet = wb.create_sheet(title=SHEET_NAME)

#-------------------------------------------------------

def writeToExcelFile():
    global ROW_INDEX
    global wb
    global FIRST_WRITE
    global VALUES

    sheet = wb[SHEET_NAME]

    # for each column, at the given row index, write the respective values
    for itr in range(VARIABLES):
        temp = chr(65 + itr)    # if itr = 0, temp = A (65 is its ascii value)
        # print("HEY!!!!!!!!!---------------")
        sheet[temp + str(ROW_INDEX)] = VALUES[itr]

    ROW_INDEX = ROW_INDEX + 1
    FIRST_WRITE = False
    wb.save(filename = PATH_TO_FILE)

#-------------------------------------------------------

def parseValues(data, meta_info):
    global ROW_INDEX
    global wb
    global FIRST_WRITE
    global VALUES

    messages_k = data["messages"]

    VALUES[0] = meta_info["portal_name"]    # portal_name
    VALUES[1] = meta_info["run_id"]    #
    VALUES[2] = meta_info["workflow_name"]
    VALUES[3] = meta_info["screen_name"]

    # file/html/help/impact/description/id/message/helpUrl/tags/timestamp
    for msg_k in messages_k:
        VALUES[4] = msg_k["extract"]
        VALUES[5] = msg_k["type"]
        VALUES[6] = msg_k["message"]

        writeToExcelFile()

#-------------------------------------------------------

def processFile(filename):
    meta_info = { }
    metas = filename.split("_")

    # metas is Empty means that it's a json api that is provided,
    # NOT a file.
    if (len(metas) == 0):
        metas = ["N/A", "N/A", "N/A", "N/A", "N/A"]

    hold = 0

    while (hold < len(metas)) and (not (metas[hold].endswith("WP") or
        metas[hold].endswith("MP") or metas[hold].endswith("CP"))):
        hold += 1

    meta_info["portal_name"] = metas[hold][-2:]  # just the last 2 letters
    meta_info["run_id"] = metas[0][-4:] + "_" + metas[1] + "_" + metas[2] + "_" + metas[3]
    meta_info["workflow_name"] = metas[hold + 1]
    hold += 2

    while (hold < len(metas)) and (not "TEDS" in metas[hold]):
         meta_info["workflow_name"] = meta_info["workflow_name"]+ "_" + metas[hold]
         hold += 1

    if (hold == len(metas)):
        print("ALERT ************** ALERT *************** ALERT")
        print("Your filename doesn't follow the specified format")
        print("Faulty File: " + filename)
        return

    meta_info["screen_name"] = metas[hold]

    if (isinstance(filename, dict)):
        parseValues(filename, meta_info)
    else:
        with open(filename) as json_data:
            data = json.load(json_data)
            parseValues(data, meta_info)

#-------------------------------------------------------

"""
def main():
    # load all JSON files in directory
    jsonFiles = glob.glob('./*.json')

    if len(jsonFiles) > 0:
        # UPDATE HEADER TITLES if NEEDED (IF NEW EXCEL FILE CREATED)
        writeToExcelFile()

    # for each file, process it
    for file_i in jsonFiles:
        processFile(file_i)


if __name__ == "__main__":
    main()
"""
