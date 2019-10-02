import json
import glob
import openpyxl
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook

ROW_INDEX = 1       # keep track of Excel row index
VARIABLES = 14
VALUES = [None] * VARIABLES
VALUES_IDX = 0
VALUES_FASTER = 0
wb = None
FIRST_WRITE = True
SHEET_NAME = "Saved Data"
PATH_TO_FILE = './www/static/saved_data_axe.xlsx'

#-------------------------------------------------------

def writeHeader():
    global wb
    global VALUES
    global ROW_INDEX

    sheet = wb[SHEET_NAME]

    VALUES[0]   = "Portal Name"
    VALUES[1]   = "Run ID"
    VALUES[2]   = "Workflow Name"
    VALUES[3]   = "Screen Name"
    VALUES[4]   = "File"
    VALUES[5]   = "Html"
    VALUES[6]   = "Help"
    VALUES[7]   = "Impact"
    VALUES[8]   = "Description"
    VALUES[9]   = "ID"
    VALUES[10]  = "Message"
    VALUES[11]  = "HelpURL"
    VALUES[12]  = "Tags"
    VALUES[13]  = "Timestamp"
    # file/html/help/impact/description/id/message/helpUrl/tags/timestamp
    for itr in range(VARIABLES):
        temp = chr(65 + itr)    # if itr = 0, temp = A (65 is its ascii value)
        sheet[temp + '1'] = VALUES[itr]

    ROW_INDEX = ROW_INDEX + 1

#------------------------------------------------------

def createNewExcelFile():
    global wb

    wb = Workbook()
    sheet = wb.create_sheet(title=SHEET_NAME)

#-------------------------------------------------------

def writeToExcelFile():
    global ROW_INDEX
    global wb
    global FIRST_WRITE
    global VALUES

    sheet = wb[SHEET_NAME]

    # for each column, at the given row index, write the respective values
    for itr in range(VARIABLES):
        temp = chr(65 + itr)    # if itr = 0, temp = A (65 is its ascii value)
        # print("HEY!!!!!!!!!---------------")
        sheet[temp + str(ROW_INDEX)] = VALUES[itr]

    ROW_INDEX = ROW_INDEX + 1
    FIRST_WRITE = False
    wb.save(filename = PATH_TO_FILE)

#-------------------------------------------------------

def parseValues(data, meta_info):
    global ROW_INDEX
    global wb
    global FIRST_WRITE
    global VALUES

    passes_k = data["passes"]
    violations_k = data["violations"]
    url_k = data["url"]
    timestamp_k = data["timestamp"]

    VALUES[0] = meta_info["portal_name"]    # portal_name
    VALUES[1] = meta_info["run_id"]    #
    VALUES[2] = meta_info["workflow_name"]
    VALUES[3] = meta_info["screen_name"]

    VALUES[4] = url_k   # filename
    VALUES[13] = timestamp_k

    # file/html/help/impact/description/id/message/helpUrl/tags/timestamp
    for violations_idx in violations_k:
        help_k = violations_idx["help"]
        description_k = violations_idx["description"]
        helpUrl_k = violations_idx["helpUrl"]
        tags_k = violations_idx["tags"]
        nodes_k = violations_idx["nodes"]

        VALUES[6] = help_k
        VALUES[8] = description_k
        VALUES[11] = helpUrl_k
        hold = ""

        for t in range(len(tags_k)):
            if t == 0:
                hold = tags_k[t]
            else:
                hold = hold + ", " + tags_k[t]

        VALUES[12] = hold
        # print(VALUES[8])
        for nodes_idx in nodes_k:

            html_k = nodes_idx["html"]
            any_k = nodes_idx["any"]      # be carefull!!! 'any' is a Python reserve word

            VALUES[5] = html_k

            for any_idx in any_k:
                impact_k = any_idx["impact"]
                id_k = any_idx["id"]
                message_k = any_idx["message"]

                VALUES[7] = impact_k
                VALUES[9] = id_k
                VALUES[10] = message_k

                writeToExcelFile()

            if (len(any_k) == 0):
                VALUES[7] = violations_idx["impact"]
                VALUES[9] = violations_idx["id"]
                VALUES[10] = "[Nothing]"

                writeToExcelFile();

#-------------------------------------------------------

def processFile(filename):
    meta_info = { }
    metas = filename.split("_")

    # metas is Empty means that it's a json api that is provided,
    # NOT a file.
    if (len(metas) == 0):
        metas = ["N/A", "N/A", "N/A", "N/A", "N/A",
            "N/A", "N/A", "N/A", "N/A", "N/A",
            "N/A", "N/A", "N/A", "N/A", "N/A",
            "N/A", "N/A", "N/A", "N/A", "N/A",
            "N/A", "N/A", "N/A", "N/A", "N/A" ]

    hold = 0

    while (hold < len(metas)) and (not ((metas[hold].endswith("WP")) or
        (metas[hold].endswith("MP")) or (metas[hold].endswith("CP"))) ):
        hold += 1

    meta_info["portal_name"] = metas[hold][-2:]  # just the last 2 letters
    meta_info["run_id"] = metas[0][-4:] + "_" + metas[1] + "_" + metas[2] + "_" + metas[3]
    meta_info["workflow_name"] = metas[hold + 1]
    hold += 2

    while (hold < len(metas)) and (not "TEDS" in metas[hold]):
         meta_info["workflow_name"] = meta_info["workflow_name"]+ "_" + metas[hold]
         hold += 1

    if (hold == len(metas)):
        print("ALERT ************** ALERT *************** ALERT")
        print("Your filename doesn't follow the specified format")
        print("Faulty File: " + filename)
        return

    meta_info["screen_name"] = metas[hold]

    if (isinstance(filename, dict)):
        parseValues(filename, meta_info)
    else:
        with open(filename) as json_data:
            data = json.load(json_data)
            parseValues(data, meta_info)

#-------------------------------------------------------

"""
def main():
    # load all JSON files in directory
    jsonFiles = glob.glob('./*.json')

    if len(jsonFiles) > 0:
        # UPDATE HEADER TITLES if NEEDED (IF NEW EXCEL FILE CREATED)
        writeToExcelFile()

    # for each file, process it
    for file_i in jsonFiles:
        processFile(file_i)


if __name__ == "__main__":
    main()
"""
